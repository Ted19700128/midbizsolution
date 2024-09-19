#!-- D:/web/midbizsolution/pfcstandard/views.py

from io import BytesIO
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.db import transaction
from .models import PFCS
from .forms import DocumentForm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def pfcs_menu(request):
    mode = request.GET.get('mode', 'view')
    show_table = True  # 항상 테이블을 표시하도록 설정
    documents = PFCS.objects.all() if show_table else None
       
    if mode == 'edit':
        document_id = request.GET.get('document_id')  # GET 파라미터에서 가져오기
        if document_id:
            document = get_object_or_404(PFCS, id=document_id)
            update_url = reverse('update_pfcs', args=[document.id])
            return redirect(update_url)
        else:
            return redirect('pfcs_menu')
    
    context = {
        'create_document': reverse('create_document'),
        'mode': mode,
        'show_table': show_table,
        'documents': documents,
    }

    return render(request, 'pfcstandard/pfcs_menu.html', context)

def show_search_popup(request):
    return render(request, 'pfcstandard/searchpopup.html')

def document_detail(request):
    document_number = request.GET.get('document_number')
    document = get_object_or_404(document, number=document_number)
    
    return render(request, 'pfcs_detail.html', {'document': document})

def create_pfcs(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST)
        if form.is_valid():
            with transaction.atomic():  # 트랜잭션으로 묶어 동시성 문제 방지
                last_document = PFCS.objects.select_for_update().order_by('id').last()
                if last_document and last_document.document_number.startswith('SD-PF01-00-'):
                    last_document_number = int(last_document.document_number.split('-')[-1])
                    new_document_number = f'SD-PF01-00-{last_document_number + 1:03d}'
                else:
                    new_document_number = 'SD-PF01-00-001'
                    
                # 새 문서 생성
                new_document = form.save(commit=False)
                new_document.document_number = new_document_number
                new_document.save()

            return redirect('pfcs_menu')  # 생성 후 문서 목록 페이지로 리다이렉트
    else:
        form = DocumentForm()
    
    return render(request, 'pfcstandard/create_pfcs.html', {'form': form})

def update_pfcs(request, document_id):
    document = get_object_or_404(PFCS, id=document_id)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            # messages.success(request, "장비가 성공적으로 업데이트되었습니다.")
            return redirect('pfcs_menu')  # 적절한 URL 이름으로 변경
        else:
            messages.error(request, "입력한 정보에 오류가 있습니다.")
            # 추가된 코드: 폼 오류를 템플릿에 전달
            return render(request, 'pfcstandard/update_pfcs.html', {'form': form, 'document': document})
    else:
        form = DocumentForm(instance=document)
    
    context = {
        'form': form,
        'document': document,
    }
    
    return render(request, 'pfcstandard/update_pfcs.html', context)

def delete_pfcs(request, document_id):
    document = get_object_or_404(PFCS, id=document_id)  # 장비가 존재하는지 확인
    
    if request.method == 'POST':
        # 'confirm_delete' 버튼이 클릭되었을 때 장비 삭제
        if 'confirm_delete' in request.POST:
            document.delete()  # 장비 삭제
            messages.success(request, "장비가 성공적으로 삭제되었습니다.")
            return redirect('document_menu')  # 삭제 후 장비 목록으로 리디렉션
        else:
            return redirect('document_menu')  # '아니오' 버튼 클릭 시 장비 목록으로 리디렉션
    
    return render(request, 'pfcstandard/delete_pfcs.html', {'documents': [document]})

def export_to_excel(request):
    filename = request.GET.get('filename', 'pfcstandard.xlsx')
    documents = PFCS.objects.all()

    # 데이터프레임 생성
    data = [
        {
        '문서 번호': documents.document_number,
        '설비 번호': documents.document_number,
        '설비명': documents.name,
        '관리부서': documents.management_team,
        '작성일자': documents.date_written,
        '설비등급': documents.rating,
        '점검주기': documents.insp_interval,
          
        '번호': documents.order,
        '점검부위': documents.insp_point,
        '점검항목': documents.insp_item,
        '등급별 점검주기': documents.insp_int_rating,
        '점검방법': documents.insp_method,
        '판정기준': documents.judge_criteria,
        '필요조치': documents.actions_required,        
        }
        for docupment in documents
    ]

    df = pd.DataFrame(data)

    # 엑셀 파일을 메모리에 생성
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Document')  # 'Document'라는 시트 이름으로 저장
    
        # 현재 워크북과 시트를 가져옴
        workbook = writer.book
        worksheet = writer.sheets['Document']  # 이제 'Document' 시트가 생성됨
        
        # 열 너비 설정
        # A열부터 AZ열까지의 너비를 2.5로 설정
        for col in range(1, 53):  # 1부터 52까지 (A부터 AZ까지)
            col_letter = worksheet.cell(row=1, column=col).column_letter
            worksheet.column_dimensions[col_letter].width = 2.5

        worksheet.row_dimensions[1].height = 15
        worksheet.row_dimensions[2].height = 25
        worksheet.row_dimensions[3].height = 25
        worksheet.row_dimensions[4].height = 15
        
        # 5행부터 100행까지의 높이를 20으로 설정
        for row in range(5, 101):
            worksheet.row_dimensions[row].height = 20
    
        # 첫 번째 행 병합 및 제목 삽입
        worksheet.merge_cells('A1:AE4')
        title_cell = worksheet['A1']
        title_cell.value = '설비 점검 기준서 (갑지)'  # 제목 설정
        title_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        title_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬

        worksheet.merge_cells('A5:F5')
        doc_num_cell = worksheet['A5']
        doc_num_cell.value = '문 서 번 호'
        doc_num_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        doc_num_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        worksheet.merge_cells('G5:K5')

        worksheet.merge_cells('L5:Q5')
        equip_num_cell = worksheet['L5']
        equip_num_cell.value = '설 비 번 호'
        equip_num_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        equip_num_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        worksheet.merge_cells('R5:W5')

        worksheet.merge_cells('X5:AC5')
        equip_name_cell = worksheet['X5']
        equip_name_cell.value = '설  비  명'
        equip_name_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        equip_name_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        worksheet.merge_cells('AD5:AI5')
        
        worksheet.merge_cells('AJ5:AO5')
        manage_team_cell = worksheet['AJ5']
        manage_team_cell.value = '관 리 부 서'
        manage_team_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        manage_team_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        worksheet.merge_cells('AP5:AU5')
    
    output.seek(0)  # 파일 포인터를 시작 위치로 이동
    
    # HttpResponse에 엑셀 파일 작성
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response