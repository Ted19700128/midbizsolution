#!-- D:/web/midbizsolution/pemledger/views.py

from io import BytesIO
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.db import transaction
from .models import Equipment
from .forms import EquipmentForm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def equipment_layout_main(request):
    return render(request, 'pemledger/equipment_layout_main.html')

def pemledger_change_main(request):
    equipments = Equipment.objects.all()
    return render(request, 'pemledger/pemledger_change_main.html', {'equipments': equipments})

def pemledger_all_main(request):
    mode = request.GET.get('mode', 'view')
    show_table = True  # 항상 테이블을 표시하도록 설정
    equipments = Equipment.objects.all() if show_table else None
       
    if mode == 'change':
        equipment_id = request.GET.get('equipment_id')  # GET 파라미터에서 가져오기
        if equipment_id:
            equipment = get_object_or_404(Equipment, id=equipment_id)
            update_url = reverse('pemledger_change_table', args=[equipment.id])
            return redirect(update_url)
        else:
            return redirect('pemledger_change_main_mode')  # 적절한 URL 이름으로 변경
    
    context = {
        'create_pemledger': reverse('create_pemledger'),
        'mode': mode,
        'show_table': show_table,
        'equipments': equipments,
    }

    return render(request, 'pemledger/pemledger_all_main.html', context)

def create_pemledger(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            # 설비 번호 자동 부여 (예: PF001 형식)
            last_equipment = Equipment.objects.order_by('id').last()
            if last_equipment:
                last_equipment_number = int(last_equipment.equipment_number[2:])
                new_equipment_number = f'PF{last_equipment_number + 1:03d}'
            else:
                new_equipment_number = 'PF001'
                    
            # 새 설비 생성
            new_equipment = form.save(commit=False)
            new_equipment.equipment_number = new_equipment_number
            new_equipment.save()

            return redirect('pemledger_all_main')  # 생성 후 설비 목록 페이지로 리다이렉트
    else:
        form = EquipmentForm()
    
    return render(request, 'pemledger/pemledger_create_table.html', {'form': form})

def change_pemledger(request, equipment_id):
    equipment = get_object_or_404(Equipment, id=equipment_id)
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            # messages.success(request, "장비가 성공적으로 업데이트되었습니다.")
            return redirect('pemledger_change_main_mode')  # 적절한 URL 이름으로 변경
        else:
            messages.error(request, "입력한 정보에 오류가 있습니다.")
            # 추가된 코드: 폼 오류를 템플릿에 전달
            return render(request, 'pemledger/pemledger_change_table.html', {'form': form, 'equipment': equipment})
    else:
        form = EquipmentForm(instance=equipment)
    
    context = {
        'form': form,
        'equipment': equipment,
    }
    
    return render(request, 'pemledger/pemledger_change_table.html', context)

def delete_pemledger(request, equipment_id):
    equipment = get_object_or_404(Equipment, id=equipment_id)  # 장비가 존재하는지 확인
    
    if request.method == 'POST':
        # 'confirm_delete' 버튼이 클릭되었을 때 장비 삭제
        if 'confirm_delete' in request.POST:
            equipment.delete()  # 장비 삭제
            messages.success(request, "장비가 성공적으로 삭제되었습니다.")
            return redirect('pemledger_change_main_mode')  # 삭제 후 장비 목록으로 리디렉션
        else:
            return redirect('pemledger_change_main_mode')  # '아니오' 버튼 클릭 시 장비 목록으로 리디렉션
    
    return render(request, 'pemledger/pemledger_change_delete.html', {'equipments': [equipment]})

def export_to_excel(request):
    filename = request.GET.get('filename', 'equipment_list.xlsx')
    equipments = Equipment.objects.all()

    # 데이터프레임 생성
    data = [
        {
            '설비번호': equipment.equipment_number,
            '설비명': equipment.name,
            '모델명': equipment.model_name,
            '제조사': equipment.manufacturer,
            '제조년월': equipment.mfg_date,
            '제조번호': equipment.mfg_number,
            '형식': equipment.equipment_type,
            '사양': equipment.specs,
            '최초설치': equipment.first_install,
            '최초양산': equipment.first_implement,
            '현 운영장소': equipment.current_operation_place,
            '관리부서': equipment.management_team,
            '오버홀': equipment.overhaul,
            '상태': equipment.current_status,
        }
        for equipment in equipments
    ]

    df = pd.DataFrame(data)

    # 엑셀 파일을 메모리에 생성
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Equipments', startrow=2)  # 'Equipments'라는 시트 이름으로 저장
    
        # 현재 워크북과 시트를 가져옴
        workbook = writer.book
        worksheet = writer.sheets['Equipments']  # 이제 'Equipments' 시트가 생성됨
        
        # 열 너비 설정
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 10
        worksheet.column_dimensions['G'].width = 10
        worksheet.column_dimensions['H'].width = 20
        worksheet.column_dimensions['I'].width = 10
        worksheet.column_dimensions['J'].width = 10
        worksheet.column_dimensions['K'].width = 10
        worksheet.column_dimensions['L'].width = 10
        worksheet.column_dimensions['M'].width = 10
        worksheet.column_dimensions['N'].width = 10
    
        # 첫 번째 행 병합 및 제목 삽입
        worksheet.merge_cells('A1:N1')  # A1부터 H1까지 병합
        title_cell = worksheet['A1']
        title_cell.value = '설비관리대장'  # 제목 설정
        title_cell.font = Font(name='맑은 고딕', size=22, bold=True)  # 폰트 설정
        title_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬

        # 1행 높이 설정
        worksheet.row_dimensions[1].height = 30

        # 2행부터 데이터가 들어가므로, 데이터 셀의 스타일 설정
        # 모든 셀 세로 중간 맞춤, H열을 제외한 모든 열 가로 중간 맞춤
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=14):
            for cell in row:
                # 세로 가운데 맞춤
                cell.alignment = Alignment(vertical='center')
                
                # H열 제외, 가로 가운데 맞춤
                if cell.column != 8:  # H열은 제외
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    output.seek(0)  # 파일 포인터를 시작 위치로 이동
    
    # HttpResponse에 엑셀 파일 작성
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def pemledger_change_main_mode(request):
    equipments = Equipment.objects.all()  # 모든 장비 목록을 가져옵니다.
    
    # 장비 선택 후 수정할 수 있도록 URL 생성 시 equipment_id 인수를 추가합니다.
    if request.method == 'POST':
        selected_id = request.POST.get('equipment_id')  # 사용자가 선택한 장비의 ID를 가져옵니다.
        if selected_id:
            return redirect('change_pemledger', equipment_id=selected_id)  # equipment_id를 전달하여 URL 생성
        else:
            messages.error(request, "수정할 장비를 선택하세요.")
    
    return render(request, 'pemledger/pemledger_change_main.html', {'equipments': equipments})

def health_check(request):
    return HttpResponse("OK", content_type="text/plain")
